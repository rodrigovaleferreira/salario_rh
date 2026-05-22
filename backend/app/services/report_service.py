# backend/app/services/report_service.py

import io
from datetime import datetime
from decimal import Decimal
from typing import List, Dict, Any

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table,
    TableStyle, HRFlowable, PageBreak,
)
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

from sqlalchemy.orm import Session
from app.models.user import User
from app.services.salary_service import SalaryService
from app.services.employee_service import EmployeeService
from app.services.diagnostic_service import DiagnosticService
from app.repositories.employee_repository import EmployeeRepository


# ── Paleta de cores corporativa ───────────────────────────────────

BRAND_BLUE    = colors.HexColor("#2455f5")
BRAND_DARK    = colors.HexColor("#111847")
SURFACE_LIGHT = colors.HexColor("#f8f9fc")
SURFACE_BORDER= colors.HexColor("#e2e6f0")
TEXT_MAIN     = colors.HexColor("#1e2540")
TEXT_MUTED    = colors.HexColor("#6b7899")
SUCCESS_COLOR = colors.HexColor("#22c55e")
WARNING_COLOR = colors.HexColor("#f59e0b")
DANGER_COLOR  = colors.HexColor("#ef4444")


def fmt_currency(value) -> str:
    if value is None:
        return "—"
    return f"R$ {float(value):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def fmt_pct(value) -> str:
    if value is None:
        return "—"
    return f"{float(value):.1f}%"


def fmt_date(value) -> str:
    if not value:
        return "—"
    if isinstance(value, str):
        return value
    return value.strftime("%d/%m/%Y")


class ReportService:

    def __init__(self, db: Session):
        self.db          = db
        self.salary_svc  = SalaryService(db)
        self.emp_svc     = EmployeeService(db)
        self.diag_svc    = DiagnosticService(db)
        self.emp_repo    = EmployeeRepository(db)

    # ── Estilos ───────────────────────────────────────────────────

    def _get_styles(self):
        base = getSampleStyleSheet()

        styles = {
            "title": ParagraphStyle(
                "ReportTitle",
                fontSize=22,
                textColor=BRAND_DARK,
                spaceAfter=6,
                fontName="Helvetica-Bold",
                alignment=TA_LEFT,
            ),
            "subtitle": ParagraphStyle(
                "ReportSubtitle",
                fontSize=10,
                textColor=TEXT_MUTED,
                spaceAfter=20,
                fontName="Helvetica",
                alignment=TA_LEFT,
            ),
            "section": ParagraphStyle(
                "Section",
                fontSize=13,
                textColor=BRAND_BLUE,
                spaceBefore=18,
                spaceAfter=8,
                fontName="Helvetica-Bold",
                alignment=TA_LEFT,
            ),
            "body": ParagraphStyle(
                "Body",
                fontSize=9,
                textColor=TEXT_MAIN,
                spaceAfter=6,
                fontName="Helvetica",
                leading=14,
            ),
            "small": ParagraphStyle(
                "Small",
                fontSize=8,
                textColor=TEXT_MUTED,
                fontName="Helvetica",
                leading=12,
            ),
            "kpi_value": ParagraphStyle(
                "KpiValue",
                fontSize=18,
                textColor=BRAND_DARK,
                fontName="Helvetica-Bold",
                alignment=TA_CENTER,
            ),
            "kpi_label": ParagraphStyle(
                "KpiLabel",
                fontSize=7,
                textColor=TEXT_MUTED,
                fontName="Helvetica",
                alignment=TA_CENTER,
                spaceAfter=4,
            ),
            "recommendation": ParagraphStyle(
                "Recommendation",
                fontSize=9,
                textColor=colors.HexColor("#1a3fd4"),
                fontName="Helvetica",
                leading=13,
                leftIndent=10,
            ),
        }
        return styles

    # ── Header/Footer ─────────────────────────────────────────────

    def _build_header_footer(self, canvas, doc):
        """Cabeçalho e rodapé em todas as páginas."""
        canvas.saveState()
        width, height = A4

        # Header — linha azul no topo
        canvas.setFillColor(BRAND_BLUE)
        canvas.rect(0, height - 1.2 * cm, width, 0.3 * cm, fill=1, stroke=0)

        # Nome do sistema
        canvas.setFont("Helvetica-Bold", 8)
        canvas.setFillColor(BRAND_BLUE)
        canvas.drawString(2 * cm, height - 0.9 * cm, "SalaryPlatform")

        # Data de geração
        canvas.setFont("Helvetica", 7)
        canvas.setFillColor(TEXT_MUTED)
        now = datetime.now().strftime("%d/%m/%Y às %H:%M")
        canvas.drawRightString(
            width - 2 * cm, height - 0.9 * cm,
            f"Gerado em {now}"
        )

        # Rodapé
        canvas.setFillColor(SURFACE_BORDER)
        canvas.rect(0, 0.8 * cm, width, 0.05 * cm, fill=1, stroke=0)

        canvas.setFont("Helvetica", 7)
        canvas.setFillColor(TEXT_MUTED)
        canvas.drawString(
            2 * cm, 0.4 * cm,
            "Confidencial — SalaryPlatform · Plataforma de Cargos e Salários"
        )
        canvas.drawRightString(
            width - 2 * cm, 0.4 * cm,
            f"Página {doc.page}"
        )

        canvas.restoreState()

    # ── KPI Box ───────────────────────────────────────────────────

    def _kpi_table(self, kpis: List[Dict]) -> Table:
        """
        Cria tabela de KPIs horizontais.
        kpis = [{"label": str, "value": str, "color": color}]
        """
        styles = self._get_styles()
        data   = [[
            Paragraph(k["value"], styles["kpi_value"])
            for k in kpis
        ], [
            Paragraph(k["label"], styles["kpi_label"])
            for k in kpis
        ]]

        col_width = (A4[0] - 4 * cm) / len(kpis)
        t = Table(data, colWidths=[col_width] * len(kpis))
        t.setStyle(TableStyle([
            ("BACKGROUND",   (0, 0), (-1, -1), SURFACE_LIGHT),
            ("ROWBACKGROUNDS",(0,0),(-1,-1), [SURFACE_LIGHT]),
            ("BOX",          (0, 0), (-1, -1), 0.5, SURFACE_BORDER),
            ("INNERGRID",    (0, 0), (-1, -1), 0.3, SURFACE_BORDER),
            ("TOPPADDING",   (0, 0), (-1, -1), 10),
            ("BOTTOMPADDING",(0, 0), (-1, -1), 10),
            ("ALIGN",        (0, 0), (-1, -1), "CENTER"),
            ("VALIGN",       (0, 0), (-1, -1), "MIDDLE"),
            ("ROUNDEDCORNERS", [4]),
        ]))
        return t

    # ── Tabela genérica ───────────────────────────────────────────

    def _data_table(
        self,
        headers: List[str],
        rows: List[List],
        col_widths: List[float] = None,
        highlight_last: bool = False,
    ) -> Table:
        """Tabela de dados estilizada."""
        header_row = [
            Paragraph(h, ParagraphStyle(
                "TH", fontSize=7, fontName="Helvetica-Bold",
                textColor=colors.white, alignment=TA_LEFT,
            ))
            for h in headers
        ]
        body_rows = []
        for i, row in enumerate(rows):
            body_rows.append([
                Paragraph(str(cell or "—"), ParagraphStyle(
                    "TD", fontSize=8, fontName="Helvetica",
                    textColor=TEXT_MAIN, alignment=TA_LEFT,
                ))
                for cell in row
            ])

        data = [header_row] + body_rows

        avail_width = A4[0] - 4 * cm
        if not col_widths:
            col_widths = [avail_width / len(headers)] * len(headers)

        t = Table(data, colWidths=col_widths, repeatRows=1)

        style_cmds = [
            ("BACKGROUND",    (0, 0), (-1, 0),  BRAND_BLUE),
            ("TEXTCOLOR",     (0, 0), (-1, 0),  colors.white),
            ("FONTNAME",      (0, 0), (-1, 0),  "Helvetica-Bold"),
            ("FONTSIZE",      (0, 0), (-1, 0),  7),
            ("TOPPADDING",    (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ("LEFTPADDING",   (0, 0), (-1, -1), 6),
            ("RIGHTPADDING",  (0, 0), (-1, -1), 6),
            ("ROWBACKGROUNDS",(0, 1), (-1, -1), [colors.white, SURFACE_LIGHT]),
            ("LINEBELOW",     (0, 0), (-1, -1), 0.3, SURFACE_BORDER),
            ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
        ]

        t.setStyle(TableStyle(style_cmds))
        return t

    # ── Relatório completo ────────────────────────────────────────

    def generate_full_report(self, current_user: User) -> bytes:
        """
        Gera relatório executivo completo em PDF.
        Inclui: resumo, análise salarial, faixas, colaboradores críticos,
        comparativo por departamento, diagnóstico e recomendações.
        """
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=2 * cm,
            leftMargin=2 * cm,
            topMargin=2.5 * cm,
            bottomMargin=2 * cm,
            title="Relatório de Cargos e Salários",
            author="SalaryPlatform",
        )

        styles  = self._get_styles()
        story   = []
        company_id = current_user.company_id

        # ── Capa ─────────────────────────────────────────────────
        story.append(Spacer(1, 1 * cm))
        story.append(Paragraph(
            "Relatório Executivo", styles["title"]
        ))
        story.append(Paragraph(
            "Cargos, Salários e Estrutura Organizacional", styles["subtitle"]
        ))
        story.append(HRFlowable(
            width="100%", thickness=2,
            color=BRAND_BLUE, spaceAfter=16,
        ))

        # ── Sumário estatístico ───────────────────────────────────
        summary = self.salary_svc.get_statistical_summary(company_id)

        if summary and "error" not in summary:
            story.append(Paragraph("1. Resumo Executivo", styles["section"]))

            kpis = [
                {
                    "label": "Total de Colaboradores",
                    "value": str(summary.get("headcount", 0)),
                },
                {
                    "label": "Folha Total Mensal",
                    "value": fmt_currency(summary.get("total_payroll")),
                },
                {
                    "label": "Salário Médio",
                    "value": fmt_currency(summary.get("mean_salary")),
                },
                {
                    "label": "Salário Mediano",
                    "value": fmt_currency(summary.get("median_salary")),
                },
            ]
            story.append(self._kpi_table(kpis))
            story.append(Spacer(1, 0.5 * cm))

            # Distribuição
            dist = summary.get("distribution", {})
            story.append(Paragraph(
                f"<b>Distribuição por faixa:</b> "
                f"{dist.get('within_band', 0)} colaboradores dentro da faixa "
                f"({dist.get('within_pct', 0):.1f}%), "
                f"{dist.get('below_band', 0)} abaixo "
                f"({dist.get('below_pct', 0):.1f}%) e "
                f"{dist.get('above_band', 0)} acima "
                f"({dist.get('above_pct', 0):.1f}%).",
                styles["body"]
            ))

            # Estatísticas descritivas
            story.append(Spacer(1, 0.3 * cm))
            stat_headers = ["Estatística", "Valor"]
            stat_rows = [
                ["Média salarial",     fmt_currency(summary.get("mean_salary"))],
                ["Mediana salarial",   fmt_currency(summary.get("median_salary"))],
                ["Desvio padrão",      fmt_currency(summary.get("std_deviation"))],
                ["Coef. de variação",  fmt_pct(summary.get("variation_coefficient"))],
                ["Percentil 25 (P25)", fmt_currency(summary.get("p25"))],
                ["Percentil 75 (P75)", fmt_currency(summary.get("p75"))],
                ["Percentil 90 (P90)", fmt_currency(summary.get("p90"))],
                ["Menor salário",      fmt_currency(summary.get("min_salary"))],
                ["Maior salário",      fmt_currency(summary.get("max_salary"))],
            ]
            story.append(self._data_table(
                stat_headers, stat_rows,
                col_widths=[8 * cm, 8 * cm],
            ))

        # ── Comparativo por departamento ──────────────────────────
        story.append(PageBreak())
        story.append(Paragraph(
            "2. Comparativo por Departamento", styles["section"]
        ))

        dept_data = self.salary_svc.get_department_comparison(company_id)
        if dept_data:
            dept_headers = [
                "Departamento", "Headcount",
                "Salário médio", "Mínimo", "Máximo",
                "vs. Média empresa",
            ]
            dept_rows = [
                [
                    d.get("department", "—"),
                    str(d.get("headcount", 0)),
                    fmt_currency(d.get("avg_salary")),
                    fmt_currency(d.get("min_salary")),
                    fmt_currency(d.get("max_salary")),
                    fmt_pct(d.get("deviation_from_company_avg")),
                ]
                for d in dept_data
            ]
            story.append(self._data_table(
                dept_headers, dept_rows,
                col_widths=[4.5*cm, 2*cm, 3*cm, 3*cm, 3*cm, 2.5*cm],
            ))
        else:
            story.append(Paragraph(
                "Sem dados de departamentos disponíveis.", styles["body"]
            ))

        # ── Colaboradores com distorção crítica ───────────────────
        story.append(PageBreak())
        story.append(Paragraph(
            "3. Colaboradores com Distorção Crítica", styles["section"]
        ))

        analysis = self.salary_svc.analyze_employees(company_id)
        critical = [r for r in analysis if r.is_critical]

        if critical:
            story.append(Paragraph(
                f"{len(critical)} colaborador(es) com desvio superior a 20% "
                f"em relação à faixa salarial.",
                styles["body"]
            ))
            story.append(Spacer(1, 0.2 * cm))

            crit_headers = [
                "Colaborador", "Salário atual",
                "Midpoint", "Compa-ratio",
                "Posição", "Desvio",
            ]
            crit_rows = [
                [
                    r.employee_name,
                    fmt_currency(r.current_salary),
                    fmt_currency(r.salary_midpoint),
                    fmt_pct(r.compa_ratio),
                    r.position_in_range,
                    fmt_pct(r.deviation_percent),
                ]
                for r in critical[:30]  # Limita 30 linhas
            ]
            story.append(self._data_table(
                crit_headers, crit_rows,
                col_widths=[4.5*cm, 2.8*cm, 2.8*cm, 2.5*cm, 2*cm, 2.4*cm],
            ))
        else:
            story.append(Paragraph(
                "✓ Nenhum colaborador com distorção crítica identificado.",
                styles["body"]
            ))

        # ── Compressão salarial ───────────────────────────────────
        compression = self.salary_svc.detect_compression(company_id)
        if compression:
            story.append(Spacer(1, 0.5 * cm))
            story.append(Paragraph(
                "4. Compressão Salarial", styles["section"]
            ))
            comp_headers = [
                "Cargo inferior", "Média",
                "Cargo superior", "Média",
                "Diferença %", "Severidade",
            ]
            comp_rows = [
                [
                    c.get("level_lower", "—"),
                    fmt_currency(c.get("avg_lower")),
                    c.get("level_upper", "—"),
                    fmt_currency(c.get("avg_upper")),
                    fmt_pct(c.get("difference_pct")),
                    c.get("severity", "—"),
                ]
                for c in compression
            ]
            story.append(self._data_table(
                comp_headers, comp_rows,
                col_widths=[3.5*cm, 2.5*cm, 3.5*cm, 2.5*cm, 2.5*cm, 2.5*cm],
            ))

        # ── Diagnóstico e recomendações ───────────────────────────
        story.append(PageBreak())
        story.append(Paragraph(
            "5. Diagnóstico Organizacional", styles["section"]
        ))

        diagnostic = self.diag_svc.generate(current_user)
        score = diagnostic.get("score", 0)
        label = diagnostic.get("health_label", "—")

        story.append(Paragraph(
            f"<b>Score de saúde organizacional: {score}/100 — {label}</b>",
            styles["body"]
        ))
        story.append(Spacer(1, 0.3 * cm))

        # Issues
        issues = diagnostic.get("issues", [])
        if issues:
            story.append(Paragraph(
                "<b>Problemas identificados:</b>", styles["body"]
            ))
            for issue in issues:
                story.append(Paragraph(
                    f"• <b>{issue.get('title')}</b>: "
                    f"{issue.get('description', '')} "
                    f"→ {issue.get('action', '')}",
                    styles["body"]
                ))

        # Warnings
        warnings = diagnostic.get("warnings", [])
        if warnings:
            story.append(Spacer(1, 0.2 * cm))
            story.append(Paragraph(
                "<b>Alertas:</b>", styles["body"]
            ))
            for w in warnings:
                story.append(Paragraph(
                    f"• {w.get('title')}: {w.get('action', '')}",
                    styles["body"]
                ))

        # Recomendações
        recs = diagnostic.get("recommendations", [])
        if recs:
            story.append(Spacer(1, 0.3 * cm))
            story.append(Paragraph(
                "<b>Recomendações:</b>", styles["body"]
            ))
            for i, rec in enumerate(recs, 1):
                story.append(Paragraph(
                    f"{i}. {rec}", styles["recommendation"]
                ))

        # ── Rodapé do relatório ───────────────────────────────────
        story.append(Spacer(1, 1 * cm))
        story.append(HRFlowable(
            width="100%", thickness=1,
            color=SURFACE_BORDER, spaceAfter=8,
        ))
        story.append(Paragraph(
            f"Relatório gerado em {datetime.now().strftime('%d/%m/%Y às %H:%M')} "
            f"pelo SalaryPlatform. Documento confidencial.",
            styles["small"]
        ))

        # Build PDF
        doc.build(
            story,
            onFirstPage=self._build_header_footer,
            onLaterPages=self._build_header_footer,
        )

        buffer.seek(0)
        return buffer.read()

    # ── Exportação XLSX ───────────────────────────────────────────

    def generate_xlsx_export(self, current_user: User) -> bytes:
        """
        Exporta dados para Excel com múltiplas abas:
        - Colaboradores
        - Faixas Salariais
        - Análise Salarial
        - Por Departamento
        """
        import openpyxl
        from openpyxl.styles import (
            Font, Fill, PatternFill, Alignment,
            Border, Side, numbers,
        )
        from openpyxl.utils import get_column_letter

        wb    = openpyxl.Workbook()
        company_id = current_user.company_id

        # Estilos base
        header_font    = Font(bold=True, color="FFFFFF", size=10)
        header_fill    = PatternFill("solid", fgColor="2455f5")
        header_align   = Alignment(horizontal="center", vertical="center")
        currency_fmt   = 'R$ #,##0.00'
        pct_fmt        = '0.0"%"'
        thin_border    = Border(
            bottom=Side(style="thin", color="E2E6F0")
        )

        def style_header_row(ws, row: int, n_cols: int):
            for col in range(1, n_cols + 1):
                cell = ws.cell(row=row, column=col)
                cell.font    = header_font
                cell.fill    = header_fill
                cell.alignment = header_align

        def auto_width(ws):
            for col in ws.columns:
                max_len = max(
                    (len(str(cell.value or "")) for cell in col), default=0
                )
                ws.column_dimensions[
                    get_column_letter(col[0].column)
                ].width = min(max_len + 4, 40)

        # ── Aba 1: Colaboradores ──────────────────────────────────
        ws1 = wb.active
        ws1.title = "Colaboradores"

        headers1 = [
            "Nome", "Matrícula", "Cargo", "Senioridade",
            "Departamento", "Centro de Custo", "Gestor",
            "Salário Atual", "Data Admissão",
        ]
        ws1.append(headers1)
        style_header_row(ws1, 1, len(headers1))

        employees, _ = self.emp_repo.list(company_id)
        for emp in employees:
            ws1.append([
                emp.name,
                emp.registration or "",
                emp.position.title if emp.position else "",
                emp.position.seniority.value if emp.position else "",
                emp.department or "",
                emp.cost_center or "",
                emp.manager_name or "",
                float(emp.current_salary),
                str(emp.hire_date) if emp.hire_date else "",
            ])
            # Formata coluna de salário
            ws1.cell(row=ws1.max_row, column=8).number_format = currency_fmt

        ws1.freeze_panes = "A2"
        auto_width(ws1)

        # ── Aba 2: Faixas Salariais ───────────────────────────────
        from app.repositories.salary_repository import SalaryRepository
        sal_repo = SalaryRepository(self.db)
        bands = sal_repo.get_all_bands(company_id)

        ws2 = wb.create_sheet("Faixas Salariais")
        headers2 = [
            "Cargo", "Mínimo", "Midpoint", "Máximo",
            "Spread %", "Mercado P50", "Moeda", "Versão",
        ]
        ws2.append(headers2)
        style_header_row(ws2, 1, len(headers2))

        for b in bands:
            ws2.append([
                b.position.title if b.position else "",
                float(b.salary_min),
                float(b.salary_midpoint),
                float(b.salary_max),
                float(b.range_spread or 0),
                float(b.market_p50) if b.market_p50 else "",
                b.currency,
                b.version,
            ])
            for col in [2, 3, 4, 6]:
                ws2.cell(row=ws2.max_row, column=col).number_format = currency_fmt
            ws2.cell(row=ws2.max_row, column=5).number_format = pct_fmt

        ws2.freeze_panes = "A2"
        auto_width(ws2)

        # ── Aba 3: Análise Salarial ───────────────────────────────
        ws3 = wb.create_sheet("Análise Salarial")
        headers3 = [
            "Colaborador", "Salário Atual", "Mínimo",
            "Midpoint", "Máximo", "Compa-ratio",
            "Posição na Faixa", "Desvio %", "Crítico",
        ]
        ws3.append(headers3)
        style_header_row(ws3, 1, len(headers3))

        analysis = self.salary_svc.analyze_employees(company_id)
        for r in analysis:
            ws3.append([
                r.employee_name,
                float(r.current_salary),
                float(r.salary_min),
                float(r.salary_midpoint),
                float(r.salary_max),
                float(r.compa_ratio),
                r.position_in_range,
                float(r.deviation_percent),
                "Sim" if r.is_critical else "Não",
            ])
            for col in [2, 3, 4, 5]:
                ws3.cell(row=ws3.max_row, column=col).number_format = currency_fmt
            for col in [6, 8]:
                ws3.cell(row=ws3.max_row, column=col).number_format = pct_fmt

        ws3.freeze_panes = "A2"
        auto_width(ws3)

        # ── Aba 4: Por Departamento ───────────────────────────────
        ws4 = wb.create_sheet("Por Departamento")
        headers4 = [
            "Departamento", "Headcount", "Salário Médio",
            "Mínimo", "Máximo", "Folha Total",
            "Desvio vs Empresa %",
        ]
        ws4.append(headers4)
        style_header_row(ws4, 1, len(headers4))

        dept_data = self.salary_svc.get_department_comparison(company_id)
        for d in dept_data:
            ws4.append([
                d.get("department", ""),
                d.get("headcount", 0),
                d.get("avg_salary", 0),
                d.get("min_salary", 0),
                d.get("max_salary", 0),
                d.get("total_payroll", 0),
                d.get("deviation_from_company_avg", 0),
            ])
            for col in [3, 4, 5, 6]:
                ws4.cell(row=ws4.max_row, column=col).number_format = currency_fmt
            ws4.cell(row=ws4.max_row, column=7).number_format = pct_fmt

        ws4.freeze_panes = "A2"
        auto_width(ws4)

        # Salva em buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.read()